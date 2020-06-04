from django.shortcuts import render, redirect, reverse
from django.http import HttpResponse, Http404, HttpResponseRedirect, HttpRequest, JsonResponse
from .models import *
from .forms import *
from tnp import keys_private
from student.models import s_registration
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordResetForm
from django.contrib import messages
import string
from django.utils import timezone as django_time
import os
import zipfile
import boto3
import boto
from wsgiref.util import FileWrapper
from io import StringIO, BytesIO
from django.http import HttpResponse
from random import *
from .decorators import *
from django.contrib.auth.decorators import login_required
from django.utils.text import slugify
from django.contrib.auth import update_session_auth_hash
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import openpyxl
from .o_f import *
from student.models import account_activation_check, student_profile, s_registration
from django.db.models.functions import Lower
import datetime
import json
from django.template.loader import render_to_string
import boto3
from django.conf import settings
import time

from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy

# Create your views here.

#def xlxs_try(request):
#    form = file_try_form()
#    s_list = []
#    if request.method == 'POST':
#        form = file_try_form(request.POST, request.FILES)
#        if form.is_valid():
#            myfile = form.cleaned_data.get('myfile')
#            #f_obj = open(myfile,'r')
#            wb_obj = openpyxl.load_workbook(myfile)
#            sheet_obj = wb_obj.active
#            cell_obj = sheet_obj.cell(row = 1, column = 1)
#            s_list = []
#            for i in range(1,int(sheet_obj.max_row)+1):
#                data_dict ={'rollno':'','branch':''}
#                data_dict['rollno']=sheet_obj.cell(row=i, column=1).value
#                data_dict['branch']=sheet_obj.cell(row=i, column=2).value
#                s_list.append(data_dict)
#            print(cell_obj.value)
#            print(sheet_obj.max_row)
#            print(sheet_obj.max_column)
#            #f_obj.close()
#    return render(request, 'excel_try.html', {'form':form,'slist':s_list})

@login_required
@manager_required
def user_xlsx(request):
    form = file_try_form()
    s_list = []
    if request.method == 'POST':
        form = file_try_form(request.POST, request.FILES)
        if form.is_valid():
            myfile = form.cleaned_data.get('myfile')
            #f_obj = open(myfile,'r')
            wb_obj = openpyxl.load_workbook(myfile)
            sheet_obj = wb_obj.active
            for i in range(1,int(sheet_obj.max_row)+1):
                e_obj = user_creator_(slugify(str(sheet_obj.cell(row=i, column=1).value)), sheet_obj.cell(row=i, column=2).value, request)
                if e_obj==None:
                    pass
                else:
                    s_list.append(e_obj)
            messages.info(request,"Users from subitted XLSX created successsfully except following.")
    return render(request, 'excel_try.html', {'form':form,'slist':s_list,'add_user':'active color-change'})

def check(request):
    #clist = company.objects.filter(name = 'cmpn8')
    c_blist = company.objects.filter(branch_allowed__name = "IT")
    #brlist = clist.name
    #print(clist)
    #print("--------------------------------------------------------")
    #print(clist[0].branch_allowed.all()[0].name)
    for c_b in c_blist:
        print(c_b.name)
    #print("--------------------------------------------------------")
    #print(type(clist))
    #cqlist = list(clist)
    #print(type(cqlist))
    #print("-------------------slugify-----------------------------")
    #str1 = "I am Yash"
    #str2 = "I am Yash-Kulshreshtha"
    #print(slugify(str1))
    #print(slugify(str2))
    #print("---------end----------slugify-----------------------------")
    #print(brlist)
    #This py file developed by Yash Kulshreshtha
#visit https://fb.com/yksmbkkr
    return render(request, 'check.html')

@login_required
@manager_required
def home(request):
    count_list = []
    
    for i in current_batch_year.objects.all():
        heading2 = 'Total placed for '+i.name+'s'
        count2 = s_registration.objects.filter(placed = True, reg_type = i).count()
        dict2 = {'h':heading2,'c':count2}
        count_list.append(dict2)
        heading ='Total companies for '+ i.name
        count = company.objects.filter(for_batch = i).count()
        dict1 = {'h':heading,'c':count}
        count_list.append(dict1)
    for i in company_grade.objects.all():
        heading = i.grade_name + ' companies'
        count = company.objects.filter(grade = i).exclude(for_batch__name = 'Internship').count()
        dict1 = {'h':heading,'c':count}
        count_list.append(dict1)
    for i in branches.objects.all():
         for j in current_batch_year.objects.all():
             heading = 'Companies for '+j.name+'s allowing '+i.name+' branch'
             count = company.objects.filter(branch_allowed__name = i, for_batch = j).count()
             dict1 = {'h':heading,'c':count}
             count_list.append(dict1)
    arg = {
        'home_nav':'active',
        'count_list' : count_list
        }
    return render(request,'new/dashboard.html',arg)

@login_required
@manager_required
def add_user(request):
        messages.error(request,'Individual user addition method is temporarily deprecated. Instead use this excel sheet upload method.')
        return redirect('management:student_db')
        if request.method == 'POST':
            form = create_user_form(request.POST)
            if form.is_valid()  :
                email_id = form.cleaned_data.get('email')
                if User.objects.filter(username = form.cleaned_data.get('username')).count()>0:
                    messages.error(request, 'User with the submitted role number is already registered.')
                    return redirect('management:add_user')
                if User.objects.filter(email = email_id).count()>0:
                    messages.error(request, 'User with the submitted email id is already registered.')
                    return redirect('management:add_user')
                if student_email_db.objects.filter(email = email_id).count()>0:
                    messages.error(request, "Student having email "+email_id+" is already in database.")
                    return redirect('management:add_user')
                elif student_email_db.objects.filter(rollno = form.cleaned_data.get('username')).count()>0:
                    messages.error(request,"Roll No "+form.cleaned_data.get('username')+" is already in database. Ask student to activate account")
                    return redirect('management:add_user')
                min_char = 12
                max_char = 20
                allchar = string.ascii_letters + string.punctuation + string.digits
                passwd = "".join(choice(allchar) for x in range(randint(min_char, max_char)))
                user = User.objects.create_user(username = str(slugify(form.cleaned_data.get('username'))).upper(), email = email_id, password=passwd)
                
                pass_form = PasswordResetForm({'email':email_id})
                if pass_form.is_valid():
                    pass_form.save(request=request, subject_template_name = 'first_pass.txt', email_template_name='first_pass_email.html', from_email="username@gmail.com", )
                else: 
                    print("Form not valid")
                
                r_obj = student_email_db(rollno = str(form.cleaned_data.get('username')).upper(), email = email_id)
                r_obj.save()
                p_obj = profile_set(user = user, check = False)
                p_obj.save()
                s_obj = account_activation_check(roll_no = str(slugify(form.cleaned_data.get('username'))).upper(), check = True, db_obj = r_obj)
                s_obj.save()
                messages.success(request, 'User '+str(form.cleaned_data.get('username')).upper()+ ' created successfully')
                return redirect('management:add_user')
        else:
            form = create_user_form()
        return render(request, 'add_user.html', {'form': form, 'add_user':'active color-change'})

@login_required
@manager_required
def create_manager(request):
     if not request.user.is_superuser:
         raise Http404("Requested page is currently unavailable.")
     if request.method == 'POST':
            form = create_manager_form(request.POST)
            if form.is_valid()  :
                email_id = form.cleaned_data.get('email')
                if User.objects.filter(username = form.cleaned_data.get('username')).count()>0:
                    messages.error(request, 'Manager with the submitted username is already registered.')
                    return redirect('management:add_manager')
                if managers.objects.filter(email = email_id).count()>0:
                    messages.error(request, 'Manager with the submitted email id is already registered.')
                    return redirect('management:add_manager')
                min_char = 12
                max_char = 20
                allchar = string.ascii_letters + string.punctuation + string.digits
                passwd = "".join(choice(allchar) for x in range(randint(min_char, max_char)))
                user = User.objects.create_user(username = form.cleaned_data.get('username'), email = email_id, password=passwd)
                manager_object = managers(user = user,email = email_id)
                manager_object.save()
                
                pass_form = PasswordResetForm({'email':email_id})
                if pass_form.is_valid():
                    pass_form.save(request=request, subject_template_name = 'first_pass.txt', email_template_name='first_pass_email.html',)
                messages.success(request, 'Manager '+form.cleaned_data.get('username')+ ' created successfully')
                return redirect('management:add_manager')
     else:
            form = create_manager_form()
     return render(request, 'new/add_manager.html', {'form': form, 'add_admin':'active'})

@login_required
@manager_required
def add_company(request):
     if request.method == 'POST':
         form = add_company_form(request.POST)
         if form.is_valid():
             c_name = form.cleaned_data.get('name')
             if company.objects.filter(name = c_name, for_batch = form.cleaned_data['for_batch']).count()>0:
                 messages.error(request, "Company with the exact same name is already registered")
                 return redirect('management:add_company')
             grade = form.cleaned_data['grade'].grade_name
             for_batch = form.cleaned_data['for_batch'].name
             if for_batch == 'Internship' and grade != 'Dream':
                 messages.error(request,"Failed to add! Company for internship must be graded as Dream company.")
                 return redirect('management:add_company')
             slug_form = form.save(commit = False)
             slug_form.slug_name = slugify(form.cleaned_data.get('name'))
             slug_form.save()
             form.save_m2m()
             messages.success(request,"Company "+form.cleaned_data.get('name')+" added successfully")
             return redirect('management:add_company')
     else:
        form = add_company_form()
     return render(request,'new/add_company.html',{'form':form, 'add_company':'active color-change'})

@login_required
def changepass(request):
    if request.method == 'POST':
        form = PasswordChangeCustomForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('management:mod_pass')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeCustomForm(request.user)
    return render(request, 'new/changepass.html', {'form':form,'mod_pass_nav':'active color-change'})

#def search_company(request, field_type = None, slug = None):
#    if request.method =='POST':
#        form = search_company_form(request.POST)
#        if form.is_valid():
#            f_name = form.cleaned_data.get('field_choice')
#            str1 = form.cleaned_data.get('slug')
#            return redirect('management:search_company_result', field_type=f_name, slug=str1)
#    else:
#        form = search_company_form()
#    f=field_type
#    s = slug
#    if f == None:
#        c_list = company.objects.all()
#        s_msg = ''
#    else:
#        try:
#            f=str(f)
#            s=str(s)
#        except:
#            raise Http404
#        if f=='name':
#            c_list = company.objects.filter(name__icontains = s)
#            s_msg = "Reults for Company Name containing "+s
#        elif f=='grad':
#            c_list = company.objects.filter(grade__grade_name = s)
#            s_msg = "Reults for Company with Grade "+s
#        elif f=='bral':
#            c_list = company.objects.filter(branch_allowed__name = s)
#            s_msg = "Reults for Company Allowing "+s
#        elif f=='clte':
#            c_list = company.objects.filter(cutoff__lte = float(s))
#            s_msg = "Reults for Company with Cut-Off less than "+s
#        elif f=='cgte':
#            c_list = company.objects.filter(cutoff__gte = float(s))
#            s_msg = "Reults for Company with Cut-Off greater than "+s
#        else:
#            raise Http404
#    if len(list(c_list))<1:
#        messages.error(request, 'No results found for the requested parameters')
#    return render(request,'search_company.html',{'clist':c_list, 'form':form, 's_comp':'active color-change', 'smsg':s_msg})

@login_required
@manager_required
def search_company(request, field_type = None, slug = None):
    if request.method =='POST':
        form = search_company_form(request.POST)
        if form.is_valid():
            f_name = form.cleaned_data.get('field_choice')
            str1 = form.cleaned_data.get('slug')
            return redirect('management:search_company_result', field_type=f_name, slug=slugify(str1))
    else:
        form = search_company_form()
    f=field_type
    s = slug
    if f == None:
        c_list = company.objects.all()
        s_msg = ''
    else:
        try:
            f=str(f)
            s=str(s)
        except:
            raise Http404
        if f=='name':
            c_list = company.objects.filter(slug_name__icontains = slugify(s))
            s_msg = "Results for Company Name containing "+s
        elif f=='grad':
            s=s.upper()
            c_list = company.objects.filter(grade__grade_name = s)
            s_msg = "Results for Company with Grade "+s
        elif f=='bral':
            br = str(slug)
            br = br.upper()
            c_list = company.objects.filter(branch_allowed__name = br)
            s_msg = "Results for Company Allowing "+br
        elif f=='clte':
            c_list = company.objects.filter(cutoff__lte = float(s))
            s_msg = "Results for Company with Cut-Off less than "+s
        elif f=='cgte':
            c_list = company.objects.filter(cutoff__gte = float(s))
            s_msg = "Results for Company with Cut-Off greater than "+s
        elif f=='batc':
            c_list = company.objects.filter(for_batch__year = int(s))
            s_msg = "Results for Companies for batch "+s
        else:
            raise Http404
    if len(list(c_list))<1:
        messages.error(request, 'No results found for the requested parameters')
    #page = request.GET.get('page', 1)
    #paginator = Paginator(c_list, 10)
    #try:
    #    c_list_page= paginator.page(page)
    #except PageNotAnInteger:
    #    c_list_page = paginator.page(1)
    #except EmptyPage:
    #    c_list_page = paginator.page(paginator.num_pages)
    return render(request,'new/search_company.html',{'clist':c_list, 'form':form, 's_comp':'active color-change', 'smsg':s_msg})

@login_required
@manager_required
def student_db_add(request):
    form = file_try_form()
    if request.method == 'POST':
        form = file_try_form(request.POST, request.FILES)
        if form.is_valid():
            myfile = form.cleaned_data.get('myfile')
            wb_obj = openpyxl.load_workbook(myfile)
            sheet_obj = wb_obj.active
            if str(sheet_obj.cell(row=1, column=1).value).strip().upper() not in ['Roll Number'.upper(), 'Roll No'.upper(), 'Roll No.'.upper()]:
                messages.error(request,'Invalid sheet format. String in cell [1,1] should be "Roll Number".')
                return redirect('management:student_db')
            if str(sheet_obj.cell(row=1, column=11).value).strip().upper() not in ['Backlogs'.upper(), 'Backlog'.upper(), 'Back'.upper()]:
                messages.error(request,'Invalid sheet format. String in cell [1,11] should be "Backlogs".')
                return redirect('management:student_db')
            for i in range(2,int(sheet_obj.max_row)+1):
                roll_no = str(sheet_obj.cell(row=i, column=1).value).strip().upper()
                email = sheet_obj.cell(row=i, column=2).value
                sem1 = str(sheet_obj.cell(row=i, column=3).value).strip().upper()
                sem2 = str(sheet_obj.cell(row=i, column=4).value).strip().upper()
                sem3 = str(sheet_obj.cell(row=i, column=5).value).strip().upper()
                sem4 = str(sheet_obj.cell(row=i, column=6).value).strip().upper()
                sem5 = str(sheet_obj.cell(row=i, column=7).value).strip().upper()
                sem6 = str(sheet_obj.cell(row=i, column=8).value).strip().upper()
                cgpa = str(sheet_obj.cell(row=i, column=9).value).strip().upper()
                batch = str(sheet_obj.cell(row=i, column=10).value).strip().upper()
                backlogs = str(sheet_obj.cell(row=i, column=11).value).strip().upper()

                if student_email_db.objects.filter(rollno = roll_no).count()>0:
                    messages.error(request,"Roll No "+roll_no+" is already in database.")
                elif student_email_db.objects.filter(email = email).count()>0:
                    messages.error(request, "Student having email "+email+" is already in database.")
                elif not (int(batch) == current_batch_year.objects.get(name = 'Placement').year or int(batch) == current_batch_year.objects.get(name = 'Internship').year):
                    messages.error(request,"Incorrect batch year for Roll No : "+roll_no)
                else:
                    sdb_obj = student_email_db(rollno = str(roll_no).upper(), email = email,backlog=backlogs,cgpa = cgpa,sem1 = sem1, sem2=sem2, sem3 = sem3, sem4 = sem4, sem5 = sem5, sem6 = sem6, batch = batch)
                    sdb_obj.save()
                    act_obj = account_activation_check(roll_no = str(roll_no).upper(), db_obj = sdb_obj)
                    act_obj.save()
    return render(request, 'new/add_db.html',{'form':form,'std_db':'active'})

@login_required
@manager_required
def edit_company(request, id=None):
    if request.method == 'POST':
            form = add_company_form(request.POST,instance = company.objects.get(id = id))
            if form.is_valid():
                grade = form.cleaned_data['grade'].grade_name
                for_batch = form.cleaned_data['for_batch'].name
                if for_batch == 'Internship' and grade != 'Dream':
                    messages.error(request,"Update Failed! Company for internship must be graded as Dream company.")
                    return redirect('management:search_company')
                form.save()
                return redirect('management:search_company')
    id = int(id)
    form = add_company_form(instance = company.objects.get(id = id))
    return render(request, 'new/add_company.html', {'s_comp':'active','form':form})

@login_required
@manager_required
def search_user(request, field_type = None, slug = None):
    if request.method =='POST':
        form = search_student_form(request.POST)
        if form.is_valid():
            f_name = form.cleaned_data.get('field_choice')
            str1 = form.cleaned_data.get('slug')
            return redirect('management:search_user_result', field_type=f_name, slug=str1)
    else:
        form = search_student_form()
    f=field_type
    s = slug
    if f == None:
        c_list = student_profile.objects.all()
        s_msg = ''
    else:
        try:
            f=str(f)
            s=str(s)
        except:
            raise Http404
        if f=='f_name':
            c_list = student_profile.objects.filter(f_name__icontains = s)
            s_msg = "Reults for Student First Name containing "+s
        elif f=='l_name':
            c_list = student_profile.objects.filter(l_name__icontains = s)
            s_msg = "Reults for Student Last Name containing "+s
        elif f=='father':
            c_list = student_profile.objects.filter(father__icontains = s)
            s_msg = "Reults for Student's Father's Name containing "+s
        elif f=='mother':
            c_list = student_profile.objects.filter(mother__icontains = s)
            s_msg = "Reults for Student's Mother's Name containing "+s
        elif f=='branch':
            c_list = student_profile.objects.filter(branch__name__icontains = slugify(s))
            s_msg = "Reults for Student of Branch "+slugify(s)
        elif f=='roll_no':
            c_list = student_profile.objects.filter(roll_no__icontains = slugify(s))
            s_msg = "Reults for Student with Cut-Off less than "+slugify(s)
        else:
            raise Http404
    if len(list(c_list))<1:
        messages.error(request, 'No results found for the requested parameters')
    #page = request.GET.get('page', 1)
    #paginator = Paginator(c_list, 10)
    #try:
    #    c_list_page= paginator.page(page)
    #except PageNotAnInteger:
    #    c_list_page = paginator.page(1)
    #except EmptyPage:
    #    c_list_page = paginator.page(paginator.num_pages)
    return render(request,'search_user.html',{'clist':c_list, 'form':form, 's_user':'active color-change', 'smsg':s_msg})

@login_required
@manager_required
def view_registrations(request,slug = None,f_choice='all'):
    if request.method =='POST':
        form = search_reg_form(request.POST)
        if form.is_valid():
            str1 = form.cleaned_data.get('slug')
            str2 = form.cleaned_data['field_choice']
            if not form.cleaned_data['batch_choice'] is None:
                str3 = form.cleaned_data['batch_choice'].name
                return redirect(reverse('management:view_reg_search',kwargs={ 'slug':str1, 'f_choice':str2})+'?batch='+str3)
            return redirect('management:view_reg_search', slug=str1,f_choice=str2)
    else:
        form = search_reg_form()
    s = slug
    if slug==None:
        c_list = s_registration.objects.all()
        s_msg = ''
    else:
        s_msg = ''
        f_msg = '(all)'
        try:
            s=str(s)
            f=str(f_choice)
        except:
            raise Http404
        c_list = s_registration.objects.filter(company__name__icontains = s)
        if f=='placed':
            c_list = c_list.filter(placed = True)
            f_msg = '(placed only)'
        elif f=='not_placed':
            c_list = c_list.filter(placed = False)
            f_msg = '(unplaced only)'
        batch_filter = request.GET.get('batch',None)
        if batch_filter == 'Internship' or batch_filter == 'Placement':
            c_list = c_list.filter(reg_type__name__icontains = batch_filter)
            f_msg = f_msg + ' for '+batch_filter.lower()+'s'
        if c_list:
            s_msg = 'Registrations '+f_msg+' for company whose name contains <strong style="color:red !important;">"'+str(slug)+'"</strong> are'
    if len(list(c_list))<1:
        messages.error(request, 'No results found for the requested parameters')
    #page = request.GET.get('page', 1)
    #paginator = Paginator(c_list, 10)
    #try:
    #    c_list_page= paginator.page(page)
    #except PageNotAnInteger:
    #    c_list_page = paginator.page(1)
    #except EmptyPage:
    #    c_list_page = paginator.page(paginator.num_pages)
    arg = {
        'v_reg':'active',
        'clist':c_list, 
        'form':form, 
        'smsg':s_msg
        }
    return render(request, 'new/view_reg.html', arg)

@login_required
@manager_required
def add_placement(request):
    if request.method=='POST':
        form = add_placement_form(request.POST)
        if form.is_valid():
            reg_id = int(form.cleaned_data.get('rid'))
            if s_registration.objects.filter(reg_id = reg_id).count() <1:
                messages.error(request, "There are no registrations with regitration number "+str(reg_id))
            else:
                s_registration.objects.filter(reg_id = reg_id).update(placed = True)
                messages.success(request,"Placement for registration id "+str(reg_id)+" added successfully")
            return redirect('management:add_placement')
    else:
        form = add_placement_form()
    return render(request, 'new/add_placement.html', {'form':form,'add_p':'active color-change'})

def trial(request):
    return render(request, 'add_placement.html')

@login_required
@manager_required
def load_user(request):
    country_id = request.GET.get('company')
    usrs = s_registration.objects.filter(company__id=country_id).order_by('reg_id')
    return render(request, 'usrs_dropdown_list_options.html', {'usrs': usrs})

@login_required
@manager_required
def add_placement_ajax(request):
    form = trial_form()
    if request.method == 'POST':
        form = trial_form(request.POST)
        if form.is_valid():
            for reg in form.cleaned_data.get('roll_no'):
                s_registration.objects.filter(reg_id = reg.reg_id).update(placed = True)
                messages.success(request,"Placement status of "+reg.user.username+' ('+reg.user.student_profile.f_name+' '+
                                 reg.user.student_profile.l_name+')'+" updated successfully")
                sns_client = boto3.client('sns', aws_access_key_id=keys_private.sns_key_id,
                                          aws_secret_access_key=keys_private.sns_key, region_name='ap-southeast-1')
                msg = 'TNP NSIT:\nPlacement Status Update\nRoll Number: '+reg.user.username+'\nPlaced In: '+reg.company.name[:20]+'\nKudos.'
                phn_number = '+91'+str(reg.user.student_profile.mobile)
                sns_response = sns_client.publish(PhoneNumber=phn_number,
                                   Message=msg,
                                   MessageAttributes={
                                       'AWS.SNS.SMS.SenderID': {'DataType': 'String', 'StringValue': 'TNPNSIT'},
                                       'AWS.SNS.SMS.SMSType': {'DataType': 'String', 'StringValue': 'Transactional'}})
                if not sns_response['ResponseMetadata'].get('HTTPStatusCode') or sns_response['ResponseMetadata'].get('HTTPStatusCode')!=200:
                    messages.warning(request,
                                     "SMS not sent to "+reg.user.username+" response - "+str(sns_response))
            return redirect('management:add_placement')
    return render(request, 'new/add_placement.html', {'form':form, 'add_p':'active'})

@login_required
@manager_required
def search_user_extended(request):
    form = search_user_extended_form()
    u_list = []
    dict_list = []
    if request.method=='POST':
        form = search_user_extended_form(request.POST)
        if form.is_valid():
            u_list = student_profile.objects.all()
            str1 = form.cleaned_data.get('slug')
            if form.cleaned_data.get('slug'):
                if str1[0].isdigit():
                    u_list = u_list.filter(roll_no__icontains = str(str1))
                else:
                    u_list = u_list.filter(f_name__icontains = str(str1))
            if form.cleaned_data.get('place'):
                check_list = s_registration.objects.filter(placed = True).values_list('user', flat=True)
                if form.cleaned_data.get('place')=='placed':
                    u_list = u_list.filter(user__id__in = check_list)
                elif form.cleaned_data.get('place')=='notplaced':
                    u_list = u_list.exclude(user__id__in = check_list)
            if form.cleaned_data.get('branch'):
                u_list = u_list.filter(branch__in = form.cleaned_data.get('branch'))
            if form.cleaned_data.get('batch'):
                u_list = u_list.filter(batch = form.cleaned_data.get('batch'))
            for u in u_list:
                dict_obj = {
                    'p':u,
                    'dream':'',
                    'app':'',
                    'ap':'',
                    'a':''
                    }
                s_reg_list = s_registration.objects.filter(user = u.user, placed = True)
                for s in s_reg_list:
                    if s.company.grade.grade == 4:
                        dict_obj['dream'] =dict_obj['dream']+ s.company.name+' --- '
                    elif s.company.grade.grade == 3:
                        dict_obj['app'] = dict_obj['app']+s.company.name+' --- '
                    elif s.company.grade.grade == 2:
                        dict_obj['ap'] = dict_obj['ap']+s.company.name+' --- '
                    elif s.company.grade.grade == 1:
                        dict_obj['a'] = dict_obj['a']+s.company.name+' --- '
                dict_list.append(dict_obj)
            return render(request,'new/search_user.html',{'form':form,'clist':dict_list,'s_user':'active'})
    return render(request,'new/search_user.html',{'form':form,'clist':dict_list,'s_user':'active'})

@login_required
@manager_required
def get_resume_zip(request):
    form = resume_download_form()
    if request.method=='POST':
        form = resume_download_form(request.POST)
        if form.is_valid():
            cmpn = form.cleaned_data.get('cmpn')
            rlist = s_registration.objects.filter(company=cmpn)
            zip_filename = "resume_collection_"+slugify(cmpn.name)+".zip"
            byte = BytesIO()
            REGION_HOST = 's3.ap-south-1.amazonaws.com'
            s3 = boto.connect_s3(settings.AWS_ACCESS_KEY_ID, settings.AWS_SECRET_ACCESS_KEY, host=REGION_HOST)
            bucket = s3.lookup(settings.AWS_STORAGE_BUCKET_NAME)
            zf = zipfile.ZipFile(byte, "w")
            #s = BytesIO()
            #zf = zipfile.ZipFile(s, "w")
            #byte = BytesIO()
            #s3 = boto3.resource('s3')
            #s3 = boto3.client('s3',aws_access_key_id=settings.AWS_ACCESS_KEY_ID, aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
            #bucket = s3.Bucket(settings.AWS_STORAGE_BUCKET_NAME)    
            #zip_file = zipfile.ZipFile(zip_filename, "w")
            zipped_files = []
            for idx,r in enumerate(rlist):
                usr = User.objects.get(id = r.user.id)
                zip_subdir = usr.username
                if r.r_file==1:
                    #usr.resume_drive.resume1.open(mode="rb")
                    fpath_file = usr.resume_drive.resume1
                    fpath = usr.resume_drive.resume1.url
                    filename = usr.resume_drive.resume1.name
                elif r.r_file==2:
                    #usr.resume_drive.resume2.open(mode="rb")
                    fpath_file = usr.resume_drive.resume2
                    fpath = usr.resume_drive.resume2.url
                    filename = usr.resume_drive.resume2.name
                elif r.r_file==3:
                    #usr.resume_drive.resume3.open(mode="rb")
                    fpath_file = usr.resume_drive.resume3
                    fpath = usr.resume_drive.resume3.url
                    filename = usr.resume_drive.resume3.name
                path = filename.split('/')
                current_file = path[len(path)-1]
                zipped_files.append(current_file)
                key = bucket.lookup(fpath.split('.com')[1])
                try:
                    data = key.read()
                except AttributeError as e:
                    # print(cmpn.name, usr.username, current_file)
                    usr.email_user(
                        subject='TNP Notification | Invalid Resume Name',
                        message='''
                        Dear {username}, <br />
                        Resume you selected for {cpn} has invalid name thus can not be included.<br />
                        Make sure, in the file name, you use Alpha Numeric characters only.<br />
                        Your selected resume - {rnum}<br />
                        Your resume's current filename - {fnm}<br />
                        Regards
                        '''.format(username=usr.username, cpn = cmpn.name, rnum = r.r_file, fnm = current_file),
                     from_email="TnP Cell - NSUT<no-reply@tnp.nsutonline.in>"
                    )
                    continue

                open(current_file, 'wb').write(data)
                zf.write(current_file,arcname=usr.username+".pdf")
                os.unlink(current_file)
                #zip_path = os.path.join(zip_subdir, filename.split('/')[1])
                #zf.write(fpath_file.read(), arcname=zip_subdir+".pdf")
                #usr.resume_drive.resume1.close()
                #usr.resume_drive.resume2.close()
                #usr.resume_drive.resume3.close()
                #path = filename.split('/')
                #current_file = path[len(path)-1]
                #key = s3.get_object(Bucket = settings.AWS_STORAGE_BUCKET_NAME, Key=fpath.split('.com')[1])
                #data = key['Body'].read()
                #open(current_file, 'wb').write(data)
                #fdir, fname = os.path.split(fpath)
                #zip_path = os.path.join(zip_subdir, path = filename.split('/')[1])
                #zf.write(current_file, zip_path)
                #os.unlink(current_file)
                #zip_file.write(fpath_file,arcname=zip_subdir+".pdf")
            #zf.close()
            zf.close()
            resp = HttpResponse(byte.getvalue(), content_type = "application/x-zip-compressed")
            resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename
            return resp
    return render(request,'new/resume_downloader.html',{'form':form,'r_download':'active color-change'})

@login_required
@manager_required
def ban_select(request):
    blist = ban.objects.all()
    form = ban_select_form()
    if request.method=='POST':
        form = ban_select_form(request.POST)
        if form.is_valid():
            rn = form.cleaned_data['rollno'].upper().strip()
            if ban.objects.filter(rollno = rn).count() > 0:
                messages.error(request,"Roll number "+rn+" is already banned. Revoke previous ban to impose new.")
                return redirect('management:ban_select')
            return redirect('management:ban_proceed',rn)
    return render(request,'new/ban_select.html',{'form':form,'ban':'active color-change','blist':blist})

@login_required
@manager_required
def ban_proceed(request, roll_no = None, ban_type = None):
    if roll_no == None:
        messages.error(request,'Invalid request parameters.')
        return redirect('management:ban_select')
    try:
        roll_no = str(roll_no)
    except:
        messages.error('Invalid request parameters.')
        return redirect('management:ban_select')
    form = ban_form()
    if request.method=='POST':
        form =ban_form(request.POST)
        if form.is_valid(): 
            if int(form.cleaned_data['field3'])<1:
                messages.error(request,"Are you planning to ban or not ? Ban for less than 1 upcoming compnay is not a ban at all.")
                return redirect('management:ban_proceed', roll_no=roll_no)
            if (form.cleaned_data['field2'] or form.cleaned_data['field1']) and form.cleaned_data['field3']:
                messages.error(request,"You can't fill first two fields if you are imposing ban for a number of upcoming companies")
                return redirect('management:ban_select')
            if not form.cleaned_data['field2'] and not form.cleaned_data['field1'] and not form.cleaned_data['field3']:
                messages.error(request,"You cant leave all fields empty.")
                return redirect('management:ban_select')
            ban_obj = ban.objects.create(rollno = roll_no, banned_by = request.user, banned_on = django_time.now())
            if form.cleaned_data['field3']:
                obj, created = ban.objects.update_or_create(rollno = roll_no, defaults={'company_count': form.cleaned_data['field3']})
            if form.cleaned_data['field1']:
                obj, created = ban.objects.update_or_create(rollno = roll_no, defaults={'till_date': form.cleaned_data['field1']})
            if form.cleaned_data['field2']:
                for i in form.cleaned_data['field2']:
                    ban_obj.companies.add(i)
            messages.success(request,'Ban is imposed on roll number '+roll_no+' !')
            return redirect('management:ban_select')
    return render(request,'new/ban_proceed.html',{'form':form,'ban':'active color-change','rollno':roll_no})

@login_required
@manager_required
def revoke_ban(request, roll_no = None):
    try:
        ban_obj = ban.objects.get(rollno = roll_no)
    except ban.DoesNotExist:
        messages.error(request,"Don't tamper with URL routing it's fool proof. Good Luck ! :P")
    ban_obj.delete()
    messages.success(request,"Ban on roll number "+roll_no+" is revoked successfuly.")
    return redirect('management:ban_select')

@login_required
@manager_required
def modify_user(request, uid = None):
    try:
        user_obj = User.objects.get(pk = uid)
    except User.DoesNotExist:
        messages.error(request,"Invalid Parameters.")
        return redirect('management:search_user')
    form = modify_user_form(instance = student_profile.objects.get(user = user_obj))
    if request.method == 'POST':
        if not request.user.is_superuser:
            messages.error(request,"You do not have permission to edit any user !")
            return redirect('management:search_user')
        form = modify_user_form(request.POST, instance = student_profile.objects.get(user = user_obj))
        if form.is_valid():
            form.save()
            messages.success(request,"Changes in user '"+user_obj.username+"' saved successfuly.")
            return redirect('management:search_user')
    return render(request,'new/modify_user.html',{'rollno':user_obj.username, 'form':form})

@login_required
@manager_required
def view_registrations(request,slug = None,f_choice='all',start_index = 0, end_index = 0):
    if request.method =='POST':
        form = search_reg_form(request.POST)
        if form.is_valid():
            str1 = form.cleaned_data.get('slug')
            str2 = form.cleaned_data['field_choice']
            if not form.cleaned_data['batch_choice'] is None:
                str3 = form.cleaned_data['batch_choice'].name
                return redirect(reverse('management:view_reg_search',kwargs={ 'slug':str1, 'f_choice':str2})+'?batch='+str3)
            return redirect('management:view_reg_search', slug=str1,f_choice=str2)
    else:
        form = search_reg_form()
    s = slug
    if slug==None:
        c_list = s_registration.objects.all()
        s_msg = ''
    else:
        s_msg = ''
        f_msg = '(all)'
        try:
            s=str(s)
            f=str(f_choice)
        except:
            raise Http404
        c_list = s_registration.objects.filter(company__name__icontains = s)
        if f=='placed':
            c_list = c_list.filter(placed = True)
            f_msg = '(placed only)'
        elif f=='not_placed':
            c_list = c_list.filter(placed = False)
            f_msg = '(unplaced only)'
        batch_filter = request.GET.get('batch',None)
        if batch_filter == 'Internship' or batch_filter == 'Placement':
            c_list = c_list.filter(reg_type__name__icontains = batch_filter)
            f_msg = f_msg + ' for '+batch_filter.lower()+'s'
        if c_list:
            s_msg = 'Registrations '+f_msg+' for company whose name contains <strong style="color:red !important;">"'+str(slug)+'"</strong> are'
    if len(list(c_list))<1:
        messages.error(request, 'No results found for the requested parameters')
    #page = request.GET.get('page', 1)
    #paginator = Paginator(c_list, 10)
    #try:
    #    c_list_page= paginator.page(page)
    #except PageNotAnInteger:
    #    c_list_page = paginator.page(1)
    #except EmptyPage:
    #    c_list_page = paginator.page(paginator.num_pages)
    arg = {
        'v_reg':'active',
        'clist':c_list,
        'form':form,
        'smsg':s_msg
        }
    return render(request, 'new/view_reg.html', arg)

@login_required
@manager_required
def view_registrations_new(request,slug = None,f_choice='all',start_index = 0, end_index = 0):
    if request.method =='POST':
        form = search_reg_form(request.POST)
        if form.is_valid():
            str1 = form.cleaned_data.get('slug')
            str2 = form.cleaned_data['field_choice']
            if not form.cleaned_data['batch_choice'] is None:
                str3 = form.cleaned_data['batch_choice'].name
                return redirect(reverse('management:view_reg_search_new',kwargs={ 'slug':str1, 'f_choice':str2})+'?batch='+str3)
            return redirect('management:view_reg_search_new', slug=str1,f_choice=str2)
    else:
        form = search_reg_form()
    s = slug
    if slug==None:
        c_list = s_registration.objects.all()[1:2]
        s_msg = 'Full list view is not supported on your browser, Search by company name.'
    else:
        s_msg = ''
        f_msg = '(all)'
        try:
            s=str(s)
            f=str(f_choice)
        except:
            raise Http404
        c_list = s_registration.objects.filter(company__name__icontains = s)
        if f=='placed':
            c_list = c_list.filter(placed = True)
            f_msg = '(placed only)'
        elif f=='not_placed':
            c_list = c_list.filter(placed = False)
            f_msg = '(unplaced only)'
        batch_filter = request.GET.get('batch',None)
        if batch_filter == 'Internship' or batch_filter == 'Placement':
            c_list = c_list.filter(reg_type__name__icontains = batch_filter)
            f_msg = f_msg + ' for '+batch_filter.lower()+'s'
        if c_list:
            s_msg = 'Registrations '+f_msg+' for company whose name contains <strong style="color:red !important;">"'+str(slug)+'"</strong> are'
    if len(list(c_list))<1:
        messages.error(request, 'No results found for the requested parameters')
    #page = request.GET.get('page', 1)
    #paginator = Paginator(c_list, 10)
    #try:
    #    c_list_page= paginator.page(page)
    #except PageNotAnInteger:
    #    c_list_page = paginator.page(1)
    #except EmptyPage:
    #    c_list_page = paginator.page(paginator.num_pages)
    c_list = list(c_list.values_list('reg_id',flat=True))
    hidden_form_ajax = view_reg_list_ajax_form()
    arg = {
        'v_reg':'active',
        'clist':c_list,
        'clist_len':len(c_list),
        'form':form,
        'hidden_form':hidden_form_ajax,
        'smsg':s_msg
        }
    return render(request, 'new/view_reg_ajax.html', arg)

@login_required
@manager_required
def get_registration_ajax(request,slug = None,f_choice='all',batch_filter=None):
    if not request.is_ajax():
        raise Http404('Invalid Request')
    if request.method == 'GET':
        raise Http404('Invalid request')
    form = view_reg_list_ajax_form(request.POST)
    if form.is_valid():
        reg_id_list = json.loads(form.cleaned_data['rList'])
        print(type(reg_id_list[0]))
        reg_obj_list = s_registration.objects.filter(reg_id__in=reg_id_list)
        html_data = render_to_string('new/reg_ajax_data.html',{'reg_obj_list':reg_obj_list})
        response_json = {'status': 200, 'html_data': html_data}
        return JsonResponse(response_json)
    response_json = {'status':400,'msg':'Oops ! Something went wrong.'}
    return JsonResponse(response_json)

@login_required
@manager_required
def batch_email_view(request):
    form = batch_email_form()
    send_email_get_email_form = batch_email_get_email_form()
    return render(request, 'new/batch_email_form.html', {'form':form,'form2':send_email_get_email_form, 'form3':batch_email_boto_form, 'send_mail_nav':'active'})

@login_required
@manager_required
def batch_email_template_preview_ajax(request):
    if request.method == 'GET' or (not request.is_ajax()):
        raise Http404('Invalid Request !')
    form = batch_email_form(request.POST)
    print(form.errors)
    if form.is_valid():
        message_body = form.cleaned_data['body']
        html_data = render_to_string('new/batch_email_template.html',{'message_body':message_body})
        return JsonResponse({'status':200,'html_data':html_data})
    return JsonResponse({'status':400,'msg':'Oops ! Something went wrong.'})

@login_required
@manager_required
def batch_email_select_ajax(request):
    form_type_flag = request.GET.get('type',None)
    if request.method == 'POST' or form_type_flag == None or (not request.is_ajax()):
        raise Http404('Invalid Request !')
    if form_type_flag == 'branch':
        form = batch_email_select_branch_form()
    elif form_type_flag == 'company':
        form = batch_email_select_company_form()
    elif form_type_flag == 'individual':
        form = batch_email_select_individual_form()
    else:
        return JsonResponse({'status':400,'html_data':'Invalid URL query parametera !'})
    html_data = render_to_string('new/batch_email_select.html', {'form':form},request=request,)
    return JsonResponse({'status':200,'html_data':html_data})

@login_required
@manager_required
def get_email_address_list(request):
    if request.method == 'GET' or (not request.is_ajax()):
        raise Http404('Invalid Request !')
    form = batch_email_get_email_form(request.POST)
    if form.is_valid():
        d1 = form.cleaned_data['json_data_field']
        if d1['flag_type'] == 'company':
            e_list = list(s_registration.objects.filter(company__id__in=d1['id_list']).values_list('user__email',flat=True).distinct())
            e_list.append('yk.smb.kkr@gmail.com')
            return JsonResponse({'status': 200, 'e_list': list(e_list)})
        elif d1['flag_type'] == 'branch':
            e_list = list(student_profile.objects.filter(branch__id__in=d1['id_list']).values_list('user__email',flat=True).distinct())
            e_list.append('yk.smb.kkr@gmail.com')
            return JsonResponse({'status': 200, 'e_list': list(e_list)})
        elif d1['flag_type'] == 'individual':
            e_list = list(User.objects.filter(id__in=d1['id_list']).values_list('email',flat=True))
            e_list.append('yk.smb.kkr@gmail.com')
            return JsonResponse({'status': 200, 'e_list': list(e_list)})
        else:
            return JsonResponse({'status': 400, 'msg': 'Invalid Request'})
    else:
        return JsonResponse({'status':400,'msg':'Invalid Request'})
    return JsonResponse({'status': 400, 'msg': 'Invalid Request'})

@login_required
@manager_required
def send_mail_boto_call_ajax(request):
    if request.method == 'GET' or (not request.is_ajax()):
        raise Http404('Invalid Request !')
    form = batch_email_boto_form(request.POST)
    if form.is_valid():
        time.sleep(1.500)
        d1 = form.cleaned_data['json_data']
        if d1['flag_type'] == 'send':
            client = boto3.client(
                'ses',
                region_name='us-east-1',
                aws_access_key_id=settings.AWS_SES_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SES_SECRET_ACCESS_KEY
            )
            destination_list = d1['id_list']
            ses_response = client.send_email(
                Source='TnP NSUT<no-reply@tnp.nsutonline.in>',
                Destination={
                    'ToAddresses': destination_list
                },
                Message={
                    'Subject': {
                        'Data': d1['e_sub'],
                        'Charset': 'UTF-8'
                    },
                    'Body': {
                        'Text': {
                            'Data': 'YOUR EMAIL CLIENT DOES NOT SUPPORT HTML EMAILS.',
                            'Charset': 'UTF-8'
                        },
                        'Html': {
                            'Data': render_to_string('new/batch_email_template.html',{'message_body':d1['e_body']}),
                            'Charset': 'UTF-8'
                        }
                    }
                },
                ReplyToAddresses=[
                    'tnpcell@nsitonline.in',
                ],
            )
            return JsonResponse({'status': 200, 'ses_response': 'Sent'})
        else:
            return JsonResponse({'status': 400, 'msg': 'Invalid Request and JSON'})
    else:
        print(form.errors)
        return JsonResponse({'status':400,'msg':'Invalid Request 2'})
    return JsonResponse({'status': 400, 'msg': 'Invalid Request'})

@login_required
@manager_required
def student_db_update(request):
    form = file_try_form()
    if request.method == 'POST':
        form = file_try_form(request.POST, request.FILES)
        if form.is_valid():
            myfile = form.cleaned_data.get('myfile')
            wb_obj = openpyxl.load_workbook(myfile)
            sheet_obj = wb_obj.active
            if str(sheet_obj.cell(row=1, column=1).value).strip().upper() not in ['Roll Number'.upper(), 'Roll No'.upper(), 'Roll No.'.upper()]:
                messages.error(request,'Invalid sheet format. String in cell [1,1] should be "Roll Number".')
                return redirect('management:student_db')
            if str(sheet_obj.cell(row=1, column=9).value).strip().upper() not in ['Backlogs'.upper(), 'Backlog'.upper(), 'Back'.upper()]:
                messages.error(request,'Invalid sheet format. String in cell [1,9] should be "Backlogs".')
                return redirect('management:student_db')
            for i in range(2,int(sheet_obj.max_row)+1):
                roll_no = str(sheet_obj.cell(row=i, column=1).value).strip().upper()
                sem1 = str(sheet_obj.cell(row=i, column=2).value).strip().upper()
                sem2 = str(sheet_obj.cell(row=i, column=3).value).strip().upper()
                sem3 = str(sheet_obj.cell(row=i, column=4).value).strip().upper()
                sem4 = str(sheet_obj.cell(row=i, column=5).value).strip().upper()
                sem5 = str(sheet_obj.cell(row=i, column=6).value).strip().upper()
                sem6 = str(sheet_obj.cell(row=i, column=7).value).strip().upper()
                cgpa = str(sheet_obj.cell(row=i, column=8).value).strip().upper()
                backlogs = str(sheet_obj.cell(row=i, column=9).value).strip().upper()

                if student_email_db.objects.filter(rollno = roll_no).count() < 1:
                    messages.error(request,"Roll No "+roll_no+" is not in database. So could not update.")
                else:
                    old_db_obj = student_email_db.objects.filter(rollno=str(roll_no)).update(backlog=backlogs,cgpa = cgpa,sem1 = sem1, sem2=sem2, sem3 = sem3, sem4 = sem4, sem5 = sem5, sem6 = sem6)
                    if student_profile.objects.filter(roll_no=roll_no).count() > 0:
                        student_profile.objects.filter(roll_no=str(roll_no)).update(backlogs=backlogs,
                                                                                                 be_marks=cgpa, sem1=sem1,
                                                                                                 sem2=sem2, sem3=sem3,
                                                                                                 sem4=sem4, sem5=sem5,
                                                                                                 sem6=sem6)

    return render(request, 'new/update_db_cgpa.html',{'form':form,'std_db_update':'active'})

@login_required
@manager_required
def add_ppo_view(request):
    form = add_ppo_form()
    if request.method == 'POST':
        form = add_ppo_form(request.POST)
        if form.is_valid():
            profile_list = form.cleaned_data['student_list']
            for profile in profile_list:
                s_registration.objects.create(reg_type=profile.batch, user=profile.user, company=form.cleaned_data['company_obj'],
                                              placed=True, r_file=1)
                messages.success(request, "PPO of '"+profile.f_name+' '+profile.l_name+"' added successfully in company '"+form.cleaned_data['company_obj'].name+"' !")
                sns_client = boto3.client('sns', aws_access_key_id=keys_private.sns_key_id,
                                          aws_secret_access_key=keys_private.sns_key, region_name='ap-southeast-1')
                msg = 'TNP NSIT:\nPPO Status Update\nRoll Number: ' + profile.user.username + '\nPlaced In: ' + form.cleaned_data['company_obj'].name[
                                                                                                                  :20] + '\nKudos.'
                phn_number = '+91' + str(profile.mobile)
                sns_response = sns_client.publish(PhoneNumber=phn_number,
                                                  Message=msg,
                                                  MessageAttributes={
                                                      'AWS.SNS.SMS.SenderID': {'DataType': 'String',
                                                                               'StringValue': 'TNPNSIT'},
                                                      'AWS.SNS.SMS.SMSType': {'DataType': 'String',
                                                                              'StringValue': 'Transactional'}})
                if not sns_response['ResponseMetadata'].get('HTTPStatusCode') or sns_response['ResponseMetadata'].get(
                        'HTTPStatusCode') != 200:
                    messages.warning(request,
                                     "SMS not sent to " + profile.user.username + " response - " + str(sns_response))
            return redirect('management:add_ppo')
    return render(request,'new/add_ppo.html',{'form':form, 'add_ppo_nav':'active'})

@login_required
@manager_required
def send_sms_view(request):
    form = sms_form()
    if request.method=='POST':
        form = sms_form(request.POST)
        if form.is_valid():
            reciepients = form.cleaned_data['receiver_list']
            message_data = form.cleaned_data['message_body']
            sender = request.user
            for profile_obj in reciepients:
                sns_client = boto3.client('sns', aws_access_key_id=keys_private.sns_key_id,
                                          aws_secret_access_key=keys_private.sns_key, region_name='ap-southeast-1')
                msg = message_data.replace('\\n','\n')
                phn_number = '+91' + str(profile_obj.mobile)
                sns_response = sns_client.publish(PhoneNumber=phn_number,
                                                  Message=msg,
                                                  MessageAttributes={
                                                      'AWS.SNS.SMS.SenderID': {'DataType': 'String',
                                                                               'StringValue': 'TNPNSIT'},
                                                      'AWS.SNS.SMS.SMSType': {'DataType': 'String',
                                                                              'StringValue': 'Transactional'}})
                if not sns_response['ResponseMetadata'].get('HTTPStatusCode') or sns_response['ResponseMetadata'].get(
                        'HTTPStatusCode') != 200:
                    messages.warning(request,
                                     "SMS not sent to " + profile_obj.user.username + " response - " + str(sns_response))
                sms_logs.objects.create(sender=sender,receiver=profile_obj.user,message_body=message_data)
            messages.success(request,"SMS Sent and logged")
            return redirect('management:send_sms')
    return render(request, 'new/send_sms.html', {'form':form, 'send_sms_nav':'active'})